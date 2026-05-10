# -*- coding: utf-8 -*-
import sys
import io
import openpyxl
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('G:/urban-ai/VP_regulation.xlsx')
result = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    result[name] = {
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'rows': rows
    }

print(json.dumps(result, ensure_ascii=False, indent=2))
